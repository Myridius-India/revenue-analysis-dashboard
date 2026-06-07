from __future__ import annotations

import sys
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from native_app.browser_dialog import CookieSnapshot
from native_app.revenue_service import load_local_dataset
from native_app.sharepoint_service import SharePointCookieRecord, build_requests_session


def build_sample_template(template_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dashboard"
    sheet["A1"] = "Customer ID"
    sheet["B1"] = "Customer Name"
    sheet["C1"] = "Project Code"
    sheet["D1"] = "Project Name"
    sheet["E1"] = "Engagement Model"
    sheet["F1"] = "Start Date"
    sheet["G1"] = "End Date"
    sheet["H1"] = "Target GM%"
    sheet.append(["C001", "Customer A", "P001", "Project Alpha", "FP", None, None, 0.25])
    workbook.save(template_path)


def build_sample_revenue_file(output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Revenue"
    headers = {
        1: "Customer Name",
        2: "Customer ID",
        3: "Project ID",
        4: "Project Name",
        7: "Engagement Model",
        8: "Start Date",
        9: "End Date",
        11: "Target GM%",
        13: "Revenue Month",
        14: "Cost Month",
        17: "Revenue YTD",
        18: "Cost YTD",
        21: "Revenue JTD",
        22: "Cost JTD",
    }
    for column, value in headers.items():
        sheet.cell(row=9, column=column, value=value)

    values = {
        1: "Customer A",
        2: "C001",
        3: "P001",
        4: "Project Alpha",
        7: "FP",
        8: "2026-01-01",
        9: "2026-12-31",
        11: 0.25,
        13: 1000,
        14: 600,
        17: 1000,
        18: 600,
        21: 1000,
        22: 600,
    }
    for column, value in values.items():
        sheet.cell(row=10, column=column, value=value)

    workbook.save(output_path)


def main() -> None:
    result_file = Path(__file__).resolve().parents[1] / "smoke_native_result.txt"
    with TemporaryDirectory() as tmp_dir:
        root = Path(tmp_dir)
        monthly_dir = root / "monthly"
        monthly_dir.mkdir()
        template_path = root / "Revenue Output Dashboard Sample.xlsx"
        revenue_path = monthly_dir / "Solutions_Revenue_January_2026.xlsx"

        build_sample_template(template_path)
        build_sample_revenue_file(revenue_path)

        dataset = load_local_dataset(monthly_dir, template_path)

        assert dataset.revenue_count == 1, dataset.revenue_count
        assert dataset.master_count == 1, dataset.master_count
        assert len(dataset.merged) == 1, len(dataset.merged)
        row = dataset.merged.iloc[0]
        assert float(row["rev_month"]) == 1000.0, row["rev_month"]
        assert float(row["cost_month"]) == 600.0, row["cost_month"]
        assert float(row["margin_month"]) == 400.0, row["margin_month"]
        assert len(dataset.snapshot) == 1, len(dataset.snapshot)
        assert len(dataset.summary) == 1, len(dataset.summary)

        print("merged:", dataset.merged[["customer_id", "project_id", "rev_month", "cost_month", "margin_month"]].to_dict("records"))
        print("snapshot:", dataset.snapshot[["customer_id", "project_id", "projected_margin"]].to_dict("records"))
        print("summary:", dataset.summary[["customer_name", "project_name", "revenue_month", "cost_month"]].to_dict("records"))

        session = build_requests_session(
            [
                SharePointCookieRecord(name="FedAuth", value="abc", domain=".sharepoint.com"),
                SharePointCookieRecord(name="rtFa", value="def", domain=".sharepoint.com"),
            ]
        )
        assert session.cookies.get("FedAuth") == "abc"
        assert session.cookies.get("rtFa") == "def"

        snapshot = CookieSnapshot(name="FedAuth", value="abc", domain=".sharepoint.com", path="/", secure=True)
        assert snapshot.name == "FedAuth"

        result_file.write_text(
            "ok\n"
            f"revenue_count={dataset.revenue_count}\n"
            f"master_count={dataset.master_count}\n"
            f"merged_rows={len(dataset.merged)}\n"
            f"snapshot_rows={len(dataset.snapshot)}\n"
            f"summary_rows={len(dataset.summary)}\n",
            encoding="utf-8",
        )


if __name__ == "__main__":
    main()
